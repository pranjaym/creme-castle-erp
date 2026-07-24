#!/usr/bin/env python3
"""
Generate deterministic, idempotent seed SQL for the spine masters (schema v2).
Reproducible: same inputs always produce the same SQL. Run from the repo root:

    python3 scripts/gen_seed_sql.py

Outputs:
    migrations/005_seed_locations.sql   locations + aliases (SupplyNote names as aliases)
    migrations/006_seed_skus.sql        46 intermediates (chef v2) with base_unit
    migrations/007_seed_par.sql         par stock
    migrations/008_seed_uom.sql         default entry-unit conversions to base

SKU source: seed_data/intermediate-sku-master-v2.xlsx. Locations: the fixed CK /
dispatch / spoke / warehouse set (SupplyNote legacy names as aliases) plus the
CC-... dark stores from seed_data/outlet-master.xlsx.
"""
import os
import openpyxl

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SEED = os.path.join(REPO, "seed_data")
MIG = os.path.join(REPO, "migrations")
EFFECTIVE_FROM = "2026-07-23"
PAR_LOCATION_CODE = "FREEZER-CK"

OMS_D2C = {"SPJ": "CC-DL-Shahpurjat", "FBD": "CC-FBD-Sector 15",
           "GN": "CC-ND-Alpha 2", "Meerut": "CC-UP-Meerut"}

SKU_CATEGORY = {"Sponge": "Sponge", "Ganache": "Ganache", "Other": "Sub-component"}
# chef log unit -> (base_unit, entry_unit, factor_to_base)
UNIT_TO_BASE = {"Pieces": ("piece", "piece", 1), "Trays": ("piece", "tray", 1),
                "Kg": ("gram", "kg", 1000)}

# The fixed canonical kitchen network (true role as name; SupplyNote/Petpooja as aliases).
# (code, name, type, region, parent_code, [ (system, ext_code, ext_name, note) ])
CANONICAL = [
    ("CK", "Central Kitchen (Noida)", "central_kitchen", "Delhi NCR", None, []),
    ("CK-SPONGE", "Sponge and Ganache Dept", "kitchen_department", "Delhi NCR", "CK",
     [("supplynote", "ND-CK", "ND-CK-Bread Dept", "sponge and ganache dept (legacy name)")]),
    ("CK-DESSERT", "Dessert Dept", "kitchen_department", "Delhi NCR", "CK",
     [("supplynote", "ND-CK", "ND-CK-Desserts Dept", "dessert dept (legacy name)")]),
    ("CK-CAKE", "Cake Dept", "kitchen_department", "Delhi NCR", "CK",
     [("supplynote", None, "Central Kitchen Noida",
       "cake dept; rename to ND-CK-Cake Dept later is just another alias, no migration")]),
    ("FREEZER-CK", "Central Kitchen Freezer", "freezer", "Delhi NCR", "CK", []),
    ("CDIS", "Central Dispatch", "central_dispatch", "Delhi NCR", None,
     [("supplynote", "CDN", "Central Dispatach-Noida", "SupplyNote misspelling"),
      ("petpooja", None, "Central Dispatch Noida", "Petpooja spelling")]),
    ("CWH", "Central Warehouse", "central_warehouse", "Delhi NCR", None,
     [("supplynote", "01", "Store Noida", "all vendor purchases land here")]),
    ("SK-ND-Sector 67", "Spoke: Noida Sector 67", "assembly_spoke", "Delhi NCR", None,
     [("supplynote", "DCCK", "SK-ND-Sector 67", None)]),
    ("SK-DL-Janakpuri", "Spoke: Janakpuri", "assembly_spoke", "Delhi NCR", None,
     [("supplynote", "JK", "SK-DL-Janakpuri", None)]),
    ("SK-GGN-Sikanderpur", "Spoke: Sikanderpur", "assembly_spoke", "Delhi NCR", None,
     [("supplynote", None, "SK-GGN-Sikanderpur", None)]),
]


def q(v):
    if v is None:
        return "null"
    s = str(v).strip()
    return "null" if s == "" else "'" + s.replace("'", "''") + "'"


def numlit(v):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return "null"
    return str(v)


def boollit(v):
    if v is None:
        return "null"
    s = str(v).strip().lower()
    return "true" if s in ("yes", "true", "y") else "false" if s in ("no", "false", "n") else "null"


def load_dark_stores():
    """Only the CC-... dark stores from the outlet master; the CK/dispatch/spoke/
    warehouse rows are handled canonically above."""
    wb = openpyxl.load_workbook(os.path.join(SEED, "outlet-master.xlsx"), data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    h = list(next(it))
    idx = {name: i for i, name in enumerate(h) if name}
    locs = {}
    for row in it:
        name = row[idx.get("Outlet Name-New")] if "Outlet Name-New" in idx else None
        if not name:
            continue
        name = str(name).strip()
        if not name.startswith("CC-"):
            continue
        locs[name] = {
            "city": (str(row[idx["City"]]).strip() if "City" in idx and row[idx["City"]] else None),
            "zrid": (str(row[idx["Zomato RID"]]).strip() if "Zomato RID" in idx and row[idx["Zomato RID"]] else None),
            "srid": (str(row[idx["Swiggy RID"]]).strip() if "Swiggy RID" in idx and row[idx["Swiggy RID"]] else None),
        }
    return locs


def region_for(city):
    if not city:
        return None
    c = city.strip().lower()
    if c in ("noida", "gurgaon", "gurugram", "delhi", "new delhi", "faridabad", "ghaziabad", "greater noida"):
        return "Delhi NCR"
    return {"meerut": "Meerut", "jaipur": "Jaipur", "chandigarh": "Chandigarh",
            "mohali": "Chandigarh", "zirakpur": "Chandigarh", "lucknow": "Lucknow"}.get(c, city.strip())


def gen_locations_sql(dark):
    L = ["-- Migration 005: SEED locations + aliases (generated, schema v2)",
         "-- Canonical CK/dispatch/spoke/warehouse set with SupplyNote legacy names as",
         "-- aliases, plus the CC-... dark stores. Regenerate: scripts/gen_seed_sql.py.", ""]
    L.append("insert into locations (code, name, type, city, region) values")
    vals = []
    for code, name, typ, region, _parent, _al in CANONICAL:
        vals.append(f"  ({q(code)}, {q(name)}, {q(typ)}::location_type, null, {q(region)})")
    for name in sorted(dark):
        d = dark[name]
        typ = "d2c_fulfillment" if name in OMS_D2C.values() else "dark_store"
        city, region = d["city"], region_for(d["city"])
        if name == "CC-UP-Meerut":                     # findings: source city wrongly New Delhi
            city, region = "Meerut", "Meerut"
        vals.append(f"  ({q(name)}, {q(name)}, {q(typ)}::location_type, {q(city)}, {q(region)})")
    L.append(",\n".join(vals))
    L.append("on conflict (code) do nothing;")
    L.append("")
    # parent links for the departments and freezer
    L.append("-- department + freezer sit under the Central Kitchen umbrella")
    for code, name, typ, region, parent, _al in CANONICAL:
        if parent:
            L.append(f"update locations set parent_id = (select id from locations where code = {q(parent)}) "
                     f"where code = {q(code)};")
    L.append("")
    # canonical aliases
    L.append("-- aliases: SupplyNote/Petpooja legacy names -> canonical location")
    for code, name, typ, region, parent, aliases in CANONICAL:
        for system, ext_code, ext_name, note in aliases:
            L.append(
                "insert into location_aliases (location_id, system, external_code, external_name, note)\n"
                f"  select id, {q(system)}::source_system, {q(ext_code)}, {q(ext_name)}, {q(note)} from locations where code = {q(code)}\n"
                "  on conflict (system, coalesce(external_code,''), coalesce(external_name,'')) do nothing;")
    L.append("")
    # dark-store aliases (console + Petpooja use the canonical CC-... string)
    L.append("-- dark stores: console + Petpooja both name the store by the canonical CC-... string")
    for name in sorted(dark):
        for system in ("dispatch_console", "petpooja"):
            L.append(
                "insert into location_aliases (location_id, system, external_name)\n"
                f"  select id, {q(system)}::source_system, {q(name)} from locations where code = {q(name)}\n"
                "  on conflict (system, coalesce(external_code,''), coalesce(external_name,'')) do nothing;")
    L.append("")
    L.append("-- OMS outlet-code aliases for the four D2C fulfillment stores")
    for oms_code, name in sorted(OMS_D2C.items()):
        L.append(
            "insert into location_aliases (location_id, system, external_code, external_name, note)\n"
            f"  select id, 'oms'::source_system, {q(oms_code)}, {q(name)}, 'D2C fulfillment store' from locations where code = {q(name)}\n"
            "  on conflict (system, coalesce(external_code,''), coalesce(external_name,'')) do nothing;")
    L.append("")
    return "\n".join(L) + "\n", len(CANONICAL) + len(dark)


def load_skus_v2():
    wb = openpyxl.load_workbook(os.path.join(SEED, "intermediate-sku-master-v2.xlsx"), data_only=True)
    ws = wb["Intermediate SKU Master v2"]
    return [list(r) for r in ws.iter_rows(min_row=4, values_only=True) if r and r[0]]


def gen_skus_sql(rows):
    L = ["-- Migration 006: SEED intermediate SKUs (generated, chef v2 + base_unit)",
         "-- Source: seed_data/intermediate-sku-master-v2.xlsx (46 rows). sku_type='intermediate'.",
         "-- base_unit for recipes; sort_order = chef's by-volume order. Regenerate: gen_seed_sql.py.", ""]
    L.append("insert into skus (code, name, sku_type, category, category_canonical, uom, base_unit, "
             "typical_qty_per_day, sort_order, to_spokes, shelf_life_days, notes) values")
    vals = []
    for i, r in enumerate(rows, start=1):
        code, name, typ, unit, qty, par, par_type, buf, spokes, shelf, note = (r + [None] * 11)[:11]
        cat = SKU_CATEGORY.get(str(typ).strip() if typ else "", "Sub-component")
        base = UNIT_TO_BASE.get(str(unit).strip() if unit else "", ("piece", "piece", 1))[0]
        vals.append(
            f"  ({q(code)}, {q(name)}, 'intermediate'::sku_type, {q(cat)}, {q(cat)}, {q(unit)}, "
            f"{q(base)}::base_unit, {numlit(qty)}, {i}, {boollit(spokes)}, {numlit(shelf)}, {q(note)})")
    L.append(",\n".join(vals))
    L.append("on conflict (code) do nothing;")
    L.append("")
    return "\n".join(L) + "\n", len(rows)


def gen_par_sql(rows):
    L = ["-- Migration 007: SEED par stocks (generated, chef v2). par_qty null for non-numeric par.", ""]
    for r in rows:
        code, name, typ, unit, qty, par, par_type, buf, spokes, shelf, note = (r + [None] * 11)[:11]
        pt = str(par_type).strip() if par_type else "fixed"
        L.append(
            "insert into par_stocks (sku_id, location_id, par_qty, par_type, effective_from, set_by)\n"
            f"  select s.id, l.id, {numlit(par)}, {q(pt)}, {q(EFFECTIVE_FROM)}, 'chef v2'\n"
            f"  from skus s, locations l where s.code = {q(code)} and l.code = {q(PAR_LOCATION_CODE)}\n"
            "  on conflict do nothing;")
    L.append("")
    return "\n".join(L) + "\n", len(rows)


def gen_uom_sql(rows):
    L = ["-- Migration 008: SEED default entry-unit conversions to base (generated).",
         "-- One default conversion per intermediate (kg->gram 1000, piece->piece 1, tray->piece 1).",
         "-- The 694 piece-unit pack conversions for raw materials are a workstream-zero task.", ""]
    for r in rows:
        code, name, typ, unit, *_ = (r + [None] * 11)[:11]
        base, entry, factor = UNIT_TO_BASE.get(str(unit).strip() if unit else "", ("piece", "piece", 1))
        L.append(
            "insert into uom_conversions (sku_id, entry_unit, factor_to_base, is_default_entry, effective_from, set_by)\n"
            f"  select id, {q(entry)}, {factor}, true, {q(EFFECTIVE_FROM)}, 'seed' from skus where code = {q(code)}\n"
            "  on conflict do nothing;")
    L.append("")
    return "\n".join(L) + "\n", len(rows)


def main():
    dark = load_dark_stores()
    rows = load_skus_v2()
    outs = [("005_seed_locations.sql", gen_locations_sql(dark)),
            ("006_seed_skus.sql", gen_skus_sql(rows)),
            ("007_seed_par.sql", gen_par_sql(rows)),
            ("008_seed_uom.sql", gen_uom_sql(rows))]
    for fn, (sql, n) in outs:
        with open(os.path.join(MIG, fn), "w") as f:
            f.write(sql)
        print(f"wrote {fn}  ({n})")
    assert outs[1][1][1] == 46, f"expected 46 intermediate SKUs, got {outs[1][1][1]}"
    print("OK: 46 intermediates, locations = canonical set + dark stores.")


if __name__ == "__main__":
    main()
