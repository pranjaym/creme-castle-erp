"""Parser for the Swiggy Daily-MTD xlsx (11 sheets, one brand, month restated
daily). Deterministic: no clock, no network. Every value becomes TEXT via one
canonical conversion so row hashes are stable across loads.

Schema and the six loader contracts: kitchen/migrations/210_swiggy_daily_mtd.sql
Plan: erp-plan/swiggy-database-plan.md, erp-plan/swiggy-dashboard-plan.md
"""
from __future__ import annotations

import hashlib
from collections import Counter
from datetime import date, datetime
from decimal import Decimal

import openpyxl

# One spec per sheet. cols maps source header -> landing column (headers not in
# the map, like the pre-Feb-2026 "Rest Name", are ignored). key is the natural
# key BEFORE dup_seq; dup_seq is appended to every key (contract 3).
# date_col is the source column that becomes business_date.
SHEETS = {
    "Sales": dict(
        shape="sales",
        cols={"dt": None, "restaurant_id": "restaurant_id", "brand_name": "brand_name",
              "area": "area", "city": "city", "orders": "orders", "gmv": "gmv"},
        date_col="dt",
        key=("business_date", "restaurant_id"),
    ),
    "Funnel": dict(
        shape="funnel",
        cols={"date": None, "restaurant_id": "restaurant_id", "brand_name": "brand_name",
              "area": "area", "city": "city", "menu_sessions": "menu_sessions",
              "cart_session": "cart_session", "payment_session": "payment_session",
              "order_session": "order_session"},
        date_col="date",
        key=("business_date", "restaurant_id"),
    ),
    "NTR-RR": dict(
        shape="ntr_rr",
        cols={"dt": None, "restaurant_id": "restaurant_id", "brand_name": "brand_name",
              "area": "area", "city": "city", "order_type": "order_type",
              "orders": "orders", "gmv": "gmv"},
        date_col="dt",
        key=("business_date", "restaurant_id", "order_type"),
        aggregate=True,          # contract 4: no stable row key in the source
    ),
    "Item Feedback and rating": dict(
        shape="item_feedback",
        cols={"dt": None, "restaurant_id": "restaurant_id", "brand_name": "brand_name",
              "city": "city", "area": "area", "order_id": "order_id",
              "gmv_total": "gmv_total", "item_name": "item_name",
              "comments": "comments", "restaurant_rating": "restaurant_rating",
              "post_status": "post_status"},
        date_col="dt",
        key=("order_id", "item_name"),
        strip_quotes=("item_name", "comments"),
    ),
    "Item sales": dict(
        shape="item_sales",
        cols={"order_id": "order_id", "ordered_time": "ordered_time",
              "restaurant_id": "restaurant_id", "brand_name": "brand_name",
              "city": "city", "area": "area", "dt": None, "item_id": "item_id",
              "item_name": "item_name", "item_category": "item_category",
              "variant_name": "variant_name", "item_quantity": "item_quantity",
              "item_subtotal": "item_subtotal", "price_per_item": "price_per_item"},
        date_col="dt",
        key=("order_id", "item_id", "variant_name", "price_per_item"),
        strip_quotes=("item_name", "item_category", "variant_name"),
    ),
    "Outlet rating": dict(
        shape="outlet_rating",
        cols={"dt": None, "restaurant_id": "restaurant_id", "brand_name": "brand_name",
              "city": "city", "area": "area", "avg_rating": "avg_rating"},
        date_col="dt",
        key=("business_date", "restaurant_id"),
    ),
    "Slot Wise Sales": dict(
        shape="slot_sales",
        cols={"dt": None, "restaurant_id": "restaurant_id", "brand_name": "brand_name",
              "city": "city", "area": "area", "order_time": "slot",
              "orders": "orders", "gmv": "gmv", "aov": "aov"},
        date_col="dt",
        key=("business_date", "restaurant_id", "slot"),
    ),
    "CPC CPV": dict(
        shape="ads_slot",
        cols={"dt": None, "time_slot": "time_slot", "rest_id": "restaurant_id",
              "city": "city", "area": "area", "brand_name": "brand_name",
              "flag": "flag", "ads_orders": "ads_orders", "ads_gmv": "ads_gmv",
              "clicks": "clicks", "budget_burnt": "budget_burnt",
              "impressions": "impressions"},
        date_col="dt",
        key=("business_date", "restaurant_id", "time_slot", "flag"),
    ),
    "Cancellation": dict(
        shape="cancellations",
        cols={"restaurant_id": "restaurant_id", "restaurant_name": "restaurant_name",
              "brand_name": "brand_name", "area": "area", "city": "city",
              "post_status": "post_status", "ordered_date": None,
              "order_id": "order_id", "ordered_time": "ordered_time",
              "is_food_prepared": "is_food_prepared", "rdc_flag": "rdc_flag",
              "cancelled_time": "cancelled_time", "item_name": "item_name",
              "cancellation_l1": "cancellation_l1", "cancellation_l2": "cancellation_l2",
              "sub_disposition_name": "sub_disposition_name"},
        date_col="ordered_date",
        key=("order_id", "item_name"),
        strip_quotes=("item_name",),
    ),
    "Coupon data": dict(
        shape="coupon_orders",
        cols={"dt": None, "order_id": "order_id",
              "restaurant_trade_discount": "restaurant_trade_discount",
              "swiggy_trade_discount": "swiggy_trade_discount",
              "restaurant_id": "restaurant_id", "brand_name": "brand_name",
              "coupon_code": "coupon_code", "coupon_discount": "coupon_discount",
              "gmv_total": "gmv_total", "swiggyit_orders": "swiggyit_orders",
              "jumbo_orders": "jumbo_orders"},
        date_col="dt",
        key=("order_id", "coupon_code"),
    ),
    "Serviceability": dict(
        shape="serviceability",
        cols={"dt": None, "restaurant_id": "restaurant_id", "brand_name": "brand_name",
              "area": "area", "city": "city", "ideal_open_hrs": "ideal_open_hrs",
              "actual_open_hrs": "actual_open_hrs"},
        date_col="dt",
        key=("business_date", "restaurant_id"),
    ),
}

# Columns that may legitimately be absent in older files (contract 5).
OPTIONAL_COLS = {"coupon_orders": {"swiggy_trade_discount"}}


def to_text(v):
    """One canonical value -> TEXT conversion, so hashes are stable.
    15-digit ids arrive as Excel ints and stringify exactly (contract 1)."""
    if v is None:
        return None
    if isinstance(v, datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0 and v.microsecond == 0:
            return v.date().isoformat()
        return v.isoformat(sep=" ")
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s if s != "" else None


def strip_wrapping_quotes(s):
    """Strip ONE balanced pair of literal double quotes (contract 2)."""
    if s is not None and len(s) >= 2 and s[0] == '"' and s[-1] == '"':
        return s[1:-1]
    return s


def to_business_date(v, sheet):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        return date.fromisoformat(v.strip()[:10])  # ISO only, never dayfirst
    raise ValueError(f"{sheet}: unparseable date cell {v!r}")


def row_hash(row: dict) -> str:
    parts = [f"{k}={row[k]}" for k in sorted(row) if not k.startswith("_")]
    return hashlib.sha256("\x1f".join(parts).encode("utf-8")).hexdigest()


def parse_file(path):
    """Returns (shapes, sheet_report) where shapes maps shape name -> list of
    row dicts (landing columns, TEXT values, business_date as date, dup_seq
    set), and sheet_report maps sheet name -> source row count (absent sheets
    are simply missing, contract 5)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    shapes, sheet_report = {}, {}
    try:
        for sheet_name, spec in SHEETS.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            it = ws.iter_rows(values_only=True)
            try:
                header = [str(h).strip() if h is not None else "" for h in next(it)]
            except StopIteration:
                continue
            colmap = spec["cols"]
            known = [h for h in header if h in colmap]
            required = {src for src, dst in colmap.items() if dst is not None}
            required -= OPTIONAL_COLS.get(spec["shape"], set())
            missing = required - set(known) - {spec["date_col"]}
            if missing:
                raise ValueError(f"{sheet_name}: missing columns {sorted(missing)}")
            if spec["date_col"] not in header:
                raise ValueError(f"{sheet_name}: missing date column {spec['date_col']}")

            rows = []
            for raw in it:
                if all(v is None for v in raw):
                    continue
                src = dict(zip(header, raw))
                out = {}
                for src_col, dst_col in colmap.items():
                    if dst_col is None or src_col not in src:
                        continue
                    val = to_text(src[src_col])
                    if val is not None and src_col in spec.get("strip_quotes", ()):
                        val = strip_wrapping_quotes(val)
                    out[dst_col] = val
                for src_col in OPTIONAL_COLS.get(spec["shape"], set()):
                    out.setdefault(colmap[src_col], None)
                out["business_date"] = to_business_date(src[spec["date_col"]], sheet_name)
                rows.append(out)
            sheet_report[sheet_name] = len(rows)

            if spec.get("aggregate"):
                rows = _aggregate_ntr_rr(rows)

            # dup_seq: 1-based occurrence of the natural key, in sheet order
            seen = Counter()
            for r in rows:
                k = tuple(str(r.get(c)) for c in spec["key"])
                seen[k] += 1
                r["dup_seq"] = seen[k]
            shapes[spec["shape"]] = rows
    finally:
        wb.close()
    return shapes, sheet_report


def _aggregate_ntr_rr(rows):
    """Contract 4: sum source rows to (business_date, restaurant_id,
    order_type). Decimal sums are exact, so the result is independent of the
    source row order and hashes stay stable across restated files."""
    agg = {}
    for r in rows:
        k = (r["business_date"], r["restaurant_id"], r["order_type"])
        a = agg.setdefault(k, {
            "business_date": r["business_date"], "restaurant_id": r["restaurant_id"],
            "order_type": r["order_type"], "brand_name": r.get("brand_name"),
            "area": r.get("area"), "city": r.get("city"),
            "_orders": Decimal(0), "_gmv": Decimal(0), "_n": 0})
        a["_orders"] += Decimal(r["orders"] or 0)
        a["_gmv"] += Decimal(r["gmv"] or 0)
        a["_n"] += 1
    out = []
    for a in agg.values():
        # format 'f' never yields exponent notation (normalize alone would
        # turn Decimal('100') into '1E+2')
        a["orders"] = format(a.pop("_orders").normalize(), "f")
        a["gmv"] = format(a.pop("_gmv").normalize(), "f")
        a["source_rows"] = a.pop("_n")
        out.append(a)
    out.sort(key=lambda r: (str(r["business_date"]), r["restaurant_id"], r["order_type"]))
    return out
